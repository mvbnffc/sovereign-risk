# Script with functions for macroeconomic modelling
from openpyxl import load_workbook
import pandas as pd
from scipy.stats import gaussian_kde
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import subprocess
import datetime
from pathlib import Path
from tqdm import tqdm
import numpy as np
import os

def update_calibration_parameters(sheet, parameter, new_value):
    '''
    This function updates the calibration sheet in DIGNAD.
    '''

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == parameter:
                # Assuming the value needs to be updated in the cell right after the parameter
                target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                try:
                    # Convert to float first
                    target_cell.value = float(new_value)
                except ValueError:
                    # if above doesn't work, save as string
                    target_cell.value = new_value
                return True  # Return after the first match to avoid unnecessary updates
    return False  # Return False if parameter not found

def update_natural_hazard_parameters(sim_start_year, nat_disaster_year, recovery_period, tradable_impact,
                                    nontradable_impact, reconstruction_efficiency, public_debt_premium,
                                    public_impact, private_impact, share_tradable):
    '''
    This function returns a dictionary that will be used to populate the Disasters sheet in DIGNAD.
    It takes as input the 9 parameters the user typically has to set in the Disasters sheet.
    '''

    excel_updates = {
    (3, 4): nat_disaster_year - sim_start_year, # C4 cell update
    (4, 4): nat_disaster_year, # D4 cell update
    (4, 7): tradable_impact, # D7 cell update
    (4, 8): nontradable_impact, # D8 cell update
    (4, 9): reconstruction_efficiency, # D9 cell update
    (4, 10): public_debt_premium, # D10 cell update
    (4, 11): public_impact, # D11 cell update
    (4, 12): private_impact, # D12 cell update
    (4, 13): share_tradable, # D13 cell update
    (3, 17): nat_disaster_year, # C17 cell update
    (4, 17): nat_disaster_year, # D17 cell update
    (3, 18): nat_disaster_year + recovery_period, # C18 cell update
    (4, 18): nat_disaster_year + recovery_period, # D18 cell update
    (3, 20): nat_disaster_year, # C20 cell update
    (3, 21): nat_disaster_year + recovery_period, # C21 cell update
    (3, 23): nat_disaster_year, # C23 cell update
    (3, 24): nat_disaster_year + recovery_period, # C24 cell update
    (3, 26): nat_disaster_year, # C26 cell update
    (3, 27): nat_disaster_year + recovery_period # C27 cell update
    }

    return excel_updates

def prepare_DIGNAD(calibration_csv, adaptation_cost, root_dir):
    '''
    This function prepares DIGNAD for being run in a simulation.
    It updates the calibration sheet with parameters from a CSV file.
    It also incorporates adaptation costs as an exogenous cost in DIGNAD. 
    Here adaptation cost is assumed to be a fixed annual investment over 5 
    years (before the disaster strikes). This is incorporated as a % of GDP.
    '''

    # Calculate adaptation cost as % of GDP
    Thai_GDP = 495.65 # in billion USD, 2022 value
    # Calculate annual adaptation cost as % of GDP (over 5 years)
    annual_adaptation_cost = (adaptation_cost / 5) / Thai_GDP * 100

    ### 1. Load the original Excel file - this is where all DIGNAD parameters are set
    # Find parent directory
    DIGNAD_root = os.path.join(root_dir, "DIGNAD", "DIGNAD_Toolkit", "DIGNAD_Toolkit")
    excel_file = os.path.join(DIGNAD_root, "input_DIG-ND.xlsx") # Assuming DIGNAD folder is in correct location
    wb = load_workbook(excel_file)

    ### 2. Load the CSV with calibration parameters
    calibration_df = pd.read_csv(calibration_csv)

    ### 3. Set calibration parameters
    sheet = wb['Calibration']
    # Iterate over the calibration DataFrame rows
    for index, row in calibration_df.iterrows():
        parameter = row['Parameters']  # The column name in your CSV for the parameter names
        new_value = row['Values']       # The column name in your CSV for the new values
        updated = update_calibration_parameters(sheet, parameter, new_value)
        if not updated:
            print(f"Parameter '{parameter}' not found in the Excel sheet.")

    ### 4. Incorporate adaptation costs
    adaptation_cost_dic = {
        (5, 4): annual_adaptation_cost,  # E4 cell update
        (5, 5): annual_adaptation_cost,  # E5 cell update
        (5, 6): annual_adaptation_cost,  # E6 cell update
        (5, 7): annual_adaptation_cost,  # E7 cell update
        (5, 8): annual_adaptation_cost   # E8 cell update
    }
    sheet = wb['Exogenous_series']
    for (col, row), value in adaptation_cost_dic.items():
        cell = sheet.cell(row=row, column=col)
        cell.value = value

    wb.save(excel_file)

def run_DIGNAD(sim_start_year, nat_disaster_year, recovery_period, tradable_impact, nontradable_impact,
                reconstruction_efficiency, public_debt_premium, public_impact, private_impact, share_tradable):
    '''
    This function runs on instance of DIGNAD with a prespecified calibration csv.
    Parameters passed to the function are the natural hazard parameters.
    Function outputs a list of GDP losses - from 2021 - 2040
    Need to set filepaths for own DIGNAD installation.
    '''

    ### 1. Set DIGNAD directory path (load Excel file)
    root = Path.cwd().parent
    DIGNAD_root = os.path.join(root, "DIGNAD", "DIGNAD_Toolkit", "DIGNAD_Toolkit")
    excel_file = os.path.join(DIGNAD_root, "input_DIG-ND.xlsx") # Assuming DIGNAD folder is in correct location
    wb = load_workbook(excel_file)

    ### 2. Update disasters sheet
    natdisaster_params = update_natural_hazard_parameters(sim_start_year, nat_disaster_year, recovery_period, tradable_impact,
                                                                nontradable_impact, reconstruction_efficiency,
                                                                public_debt_premium, public_impact, private_impact, share_tradable)
    sheet = wb['Disasters']
    for (col, row), value in natdisaster_params.items():
        cell = sheet.cell(row=row, column=col)
        cell.value = value

    ### 3. Save Excel Workbook
    wb.save(excel_file)

    ### 4. Run Matlab (note: Matlab should be added to PATH)
    matlab_script = os.path.join(DIGNAD_root, "simulate.m")
    result = subprocess.call(["matlab", "-batch", "run('" + matlab_script + "')"])
    if int(result) != 0:
        print("MATLAB script not executed succesfully")
        return None, None

    ### 5. Read results from Excel sheet
    # Get today's date as that is the name of file and directory
    today = datetime.datetime.today().strftime("%d%b%Y")
    file_path = os.path.join(DIGNAD_root, "Excel output", f"{today}", f"Model_output_{today}.xlsx")
    df = pd.read_excel(file_path)
    years = list(df.iloc[:, 0])
    gdp_impact = list(df.iloc[:, 1])

    return gdp_impact, years

def run_flood_sim_for_macro(basin_curves, adaptation_aep, num_sims, copula_random_ns, agr_GVA, man_GVA, ser_GVA, tradable_shares, thai_gdp):
    '''
    This function runs the flood simulation for macroeconomic modelling.
    It takes as input:
        - basin_curves: dict of BasinFloodDamageCurve objects, keyed by basin ID
        - adaptation_aep: float, AEP level for adaptation measures
        - num_sims: int, number of simulation years
        - copula_random_ns: pd.DataFrame, random numbers for each basin and simulation year
        - agr_GVA: float, Agriculture GVA baseline
        - man_GVA: float, Manufacturing GVA baseline
        - ser_GVA: float, Service GVA baseline
        - tradables_shares: dict, tradable shares per sector
        - thai_gdp: float, total GDP of Thailand
    It outputs two dataframes with time series of DIGNAD input metrics (baseline and adapted).
    '''
    
    # Extract all sectors from the basin curves
    all_sectors = {comp.sector for curve in basin_curves.values() for comp in curve.components}
    # Extract all basin IDs
    basin_ids = list(basin_curves.keys())
    # Initialize outputs 
    # per-sector annual losses (monetary)
    sector_baseline_losses = {s: np.zeros(num_sims) for s in all_sectors}
    sector_adapted_losses  = {s: np.zeros(num_sims) for s in all_sectors}
    # DIGNAD aggregate series (per year)
    trad_output_loss_pct_baseline    = np.zeros(num_sims)
    trad_output_loss_pct_adapted     = np.zeros(num_sims)
    nontrad_output_loss_pct_baseline = np.zeros(num_sims)
    nontrad_output_loss_pct_adapted  = np.zeros(num_sims)
    private_cap_damage_pct_gdp_baseline  = np.zeros(num_sims)
    private_cap_damage_pct_gdp_adapted   = np.zeros(num_sims)
    public_cap_damage_pct_gdp_baseline   = np.zeros(num_sims)
    public_cap_damage_pct_gdp_adapted    = np.zeros(num_sims)
    # Set some baslines
    tradable_output_baseline = (
        agr_GVA * tradable_shares["Agriculture"] +
        man_GVA * tradable_shares["Manufacturing"] +
        ser_GVA * tradable_shares["Service"]
    )
    nontrad_output_baseline = (
        agr_GVA * (1 - tradable_shares["Agriculture"]) +
        man_GVA * (1 - tradable_shares["Manufacturing"]) +
        ser_GVA * (1 - tradable_shares["Service"])
    )

    for t in tqdm(range(num_sims)):
        # sector totals for this simulated year
        sector_year_baseline = {s: 0.0 for s in all_sectors}
        sector_year_adapted  = {s: 0.0 for s in all_sectors}
        random_ns =  copula_random_ns.loc[t] # extract random numbers for year t

        for basin_id in basin_ids:
            # TEMP DEBUG (need to re-run copulas as some basins not included)
            basin_str = str(int(basin_id))
            if basin_str not in random_ns:
                continue # skip
            curve = basin_curves[basin_id]
            aep_event = 1-random_ns[str(int(basin_id))]

            for s in all_sectors:
                # baseline
                bl = curve.loss_at_event_aep(
                    aep_event,
                    scenario="baseline",
                    sector=s,
                )
                # adaptation
                ad = curve.loss_at_event_aep(
                    aep_event,
                    scenario="adaptation",
                    adapted_protection_aep=adaptation_aep,
                    sector=s,
                )
                sector_year_baseline[s] += bl
                sector_year_adapted[s]  += ad

        # store per-sector time series
        for s in all_sectors:
            sector_baseline_losses[s][t] = sector_year_baseline[s]
            sector_adapted_losses[s][t]  = sector_year_adapted[s]

        # ---- MAP TO DIGNAD AGGREGATES (BASELINE) ----
        ag_b   = sector_year_baseline.get("Agriculture", 0.0)
        man_b  = sector_year_baseline.get("Manufacturing", 0.0)
        serv_b = sector_year_baseline.get("Service", 0.0)
        priv_b = sector_year_baseline.get("Private", 0.0)
        pub_b  = sector_year_baseline.get("Public", 0.0)

        trad_out_b = (
            ag_b  * tradable_shares["Agriculture"] +
            man_b * tradable_shares["Manufacturing"] +
            serv_b* tradable_shares["Service"]
        )
        nontrad_out_b = (
            ag_b  * (1 - tradable_shares["Agriculture"]) +
            man_b * (1 - tradable_shares["Manufacturing"]) +
            serv_b* (1 - tradable_shares["Service"])
        )

        trad_output_loss_pct_baseline[t]    = trad_out_b    / tradable_output_baseline
        nontrad_output_loss_pct_baseline[t] = nontrad_out_b / nontrad_output_baseline
        private_cap_damage_pct_gdp_baseline[t]  = priv_b / thai_gdp
        public_cap_damage_pct_gdp_baseline[t]   = pub_b  / thai_gdp

        # ---- MAP TO DIGNAD AGGREGATES (ADAPTED) ----
        ag_a   = sector_year_adapted.get("Agriculture", 0.0)
        man_a  = sector_year_adapted.get("Manufacturing", 0.0)
        serv_a = sector_year_adapted.get("Service", 0.0)
        priv_a = sector_year_adapted.get("Private", 0.0)
        pub_a  = sector_year_adapted.get("Public", 0.0)

        trad_out_a = (
            ag_a  * tradable_shares["Agriculture"] +
            man_a * tradable_shares["Manufacturing"] +
            serv_a* tradable_shares["Service"]
        )
        nontrad_out_a = (
            ag_a  * (1 - tradable_shares["Agriculture"]) +
            man_a * (1 - tradable_shares["Manufacturing"]) +
            serv_a* (1 - tradable_shares["Service"])
        )

        trad_output_loss_pct_adapted[t]    = trad_out_a    / tradable_output_baseline
        nontrad_output_loss_pct_adapted[t] = nontrad_out_a / nontrad_output_baseline
        private_cap_damage_pct_gdp_adapted[t]  = priv_a / thai_gdp
        public_cap_damage_pct_gdp_adapted[t]   = pub_a  / thai_gdp  
    
    # Prepare final DataFrames
    baseline_shocks = pd.DataFrame({
        "year_index": np.arange(num_sims),
        "dY_T": trad_output_loss_pct_baseline,       # tradable output loss (% trad output)
        "dY_N": nontrad_output_loss_pct_baseline,    # non-trad output loss (% non-trad output)
        "dK_priv": private_cap_damage_pct_gdp_baseline,  # % GDP
        "dK_pub":  public_cap_damage_pct_gdp_baseline,   # % GDP
    })
    adapted_shocks = pd.DataFrame({
        "year_index": np.arange(num_sims),
        "dY_T": trad_output_loss_pct_adapted,       # tradable output loss (% trad output)
        "dY_N": nontrad_output_loss_pct_adapted,    # non-trad output loss (% non-trad output)
        "dK_priv": private_cap_damage_pct_gdp_adapted,  # % GDP
        "dK_pub":  public_cap_damage_pct_gdp_adapted,   # % GDP
    })

    return baseline_shocks, adapted_shocks

def create_dignad_parameter_grid(flood_df, n_samples=500, method='combined', 
                                 ensure_extremes=True, random_state=42):
    """
    Create parameter grid for DIGNAD pre-computation from flood simulations
    
    Parameters:
    -----------
    flood_df : pd.DataFrame
        Combined flood simulation results with [dY_T, dY_N, dK_priv, dK_pub]
    n_samples : int
        Target number of parameter combinations
    method : str
        'clustering', 'stratified', 'kde', or 'combined'
    ensure_extremes : bool
        Whether to ensure extreme events are included
    random_state : int
        Random seed for reproducibility
    
    Returns:
    --------
    param_grid : pd.DataFrame
        Parameter grid for DIGNAD simulations
    """
    np.random.seed(random_state)
    param_cols = ['dY_T', 'dY_N', 'dK_priv', 'dK_pub']
    X = flood_df[param_cols].values
    
    print(f"\nCreating DIGNAD parameter grid using '{method}' method...")
    print(f"Target samples: {n_samples}")
    
    if method == 'clustering':
        grid = _create_clustering_grid(X, param_cols, n_samples)
    
    elif method == 'stratified':
        grid = _create_stratified_grid(flood_df, param_cols, n_samples)
    
    elif method == 'kde':
        grid = _create_kde_grid(X, param_cols, n_samples)
    
    elif method == 'combined':
        # Use multiple methods for better coverage
        n_per_method = n_samples // 3
        
        grid1 = _create_clustering_grid(X, param_cols, n_per_method)
        grid2 = _create_stratified_grid(flood_df, param_cols, n_per_method)
        grid3 = _create_kde_grid(X, param_cols, n_samples - 2*n_per_method)
        
        grid = pd.concat([grid1, grid2, grid3], ignore_index=True)
        
        # Remove duplicates
        grid = grid.drop_duplicates(subset=param_cols).reset_index(drop=True)
    
    else:
        raise ValueError(f"Unknown method: {method}")
    
    # Add extreme events if requested
    if ensure_extremes:
        grid = _add_extreme_events(grid, flood_df, param_cols)
    
    # Add metadata
    grid['grid_method'] = method
    grid['grid_index'] = range(len(grid))
    
    # Rename columns for DIGNAD compatibility
    grid_renamed = grid.copy()
    grid_renamed.rename(columns={
        'dY_T': 'tradable_impact',
        'dY_N': 'nontradable_impact', 
        'dK_priv': 'private_impact',
        'dK_pub': 'public_impact'
    }, inplace=True)
    
    # Add fixed DIGNAD parameters
    grid_renamed['share_tradable'] = 0.5  # Adjust as needed
    grid_renamed['reconstruction_efficiency'] = 0
    grid_renamed['public_debt_premium'] = 0
    
    print(f"\nFinal parameter grid:")
    print(f"  Grid size: {len(grid_renamed)} unique combinations")
    print(f"  Parameter ranges:")
    for col in param_cols:
        print(f"    {col}: [{grid[col].min():.3f}, {grid[col].max():.3f}]")
    
    return grid_renamed


def _create_clustering_grid(X, param_cols, n_samples):
    """Create grid using k-means clustering"""
    
    # Standardize for clustering
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    
    # Perform k-means
    n_clusters = min(n_samples, len(X))
    kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
    kmeans.fit(X_scaled)
    
    # Get cluster centers
    centers_scaled = kmeans.cluster_centers_
    centers = scaler.inverse_transform(centers_scaled)
    
    # Also include some boundary points from each cluster
    grid_points = list(centers)
    
    for i in range(n_clusters):
        cluster_mask = kmeans.labels_ == i
        if cluster_mask.sum() > 1:
            cluster_points = X[cluster_mask]
            # Add point furthest from center
            distances = np.linalg.norm(cluster_points - centers[i], axis=1)
            if len(distances) > 0:
                furthest_idx = np.argmax(distances)
                grid_points.append(cluster_points[furthest_idx])
    
    grid = pd.DataFrame(grid_points[:n_samples], columns=param_cols)
    return grid


def _create_stratified_grid(flood_df, param_cols, n_samples):
    """Create grid using stratified sampling based on severity"""
    
    # Calculate total loss for stratification
    flood_df = flood_df.copy()
    flood_df['total_loss'] = flood_df[param_cols].sum(axis=1)
    
    # Create severity bins
    n_bins = min(20, len(flood_df) // 10)
    flood_df['severity_bin'] = pd.qcut(flood_df['total_loss'], q=n_bins, 
                                       labels=False, duplicates='drop')
    
    # Sample from each bin
    samples_per_bin = n_samples // n_bins
    remainder = n_samples % n_bins
    
    stratified_samples = []
    
    for bin_id in range(n_bins):
        bin_data = flood_df[flood_df['severity_bin'] == bin_id]
        if len(bin_data) > 0:
            # Add extra sample to some bins to reach exact n_samples
            extra = 1 if bin_id < remainder else 0
            sample_size = min(samples_per_bin + extra, len(bin_data))
            
            bin_sample = bin_data[param_cols].sample(n=sample_size, replace=False)
            stratified_samples.append(bin_sample)
    
    grid = pd.concat(stratified_samples, ignore_index=True)
    return grid


def _create_kde_grid(X, param_cols, n_samples):
    """Create grid by sampling from kernel density estimate"""
    
    # Fit KDE to joint distribution
    kde = gaussian_kde(X.T, bw_method='scott')
    
    # Sample from KDE
    samples = kde.resample(n_samples).T
    
    # Ensure samples are within data bounds
    for i in range(samples.shape[1]):
        samples[:, i] = np.clip(samples[:, i], 
                               X[:, i].min() * 0.95,  # Allow slight extrapolation
                               X[:, i].max() * 1.05)
    
    grid = pd.DataFrame(samples, columns=param_cols)
    return grid


def _add_extreme_events(grid, flood_df, param_cols):
    """Add extreme parameter combinations to ensure tail coverage"""
    
    extreme_points = []
    X = flood_df[param_cols].values
    
    # Add maximum for each parameter
    for i, col in enumerate(param_cols):
        max_idx = np.argmax(X[:, i])
        extreme_points.append(X[max_idx])
        
        # Also add 99.5th percentile
        p995_val = np.percentile(X[:, i], 99.5)
        close_idx = np.argmin(np.abs(X[:, i] - p995_val))
        extreme_points.append(X[close_idx])
    
    # Add joint extremes (high total loss)
    total_loss = X.sum(axis=1)
    extreme_indices = np.argsort(total_loss)[-20:]  # Top 20 most severe
    for idx in extreme_indices:
        extreme_points.append(X[idx])
    
    # Add to grid
    extreme_df = pd.DataFrame(extreme_points, columns=param_cols)
    grid = pd.concat([grid, extreme_df], ignore_index=True)
    grid = grid.drop_duplicates(subset=param_cols).reset_index(drop=True)
    
    return grid
